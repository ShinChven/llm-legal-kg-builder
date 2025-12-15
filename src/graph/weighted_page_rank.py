"""
Compute PageRank using Neo4j Graph Data Science (GDS) and export results.

This replaces the local Python PageRank with native GDS calls:
  - Projects an in-memory graph of `Act` nodes and selected relationships.
  - Runs PageRank in stream mode for weight sets:
    - "unweighted" (uniform weights across types)
    - One additional weighted variant (distinct notion of influence)
  - Joins node properties (title, status) and writes to Excel.
  - Optional: embeds thesis-ready charts in the same Excel (no external files).

Incoming relation composition (for context): CITES ≈ 50.97%, AMENDS ≈ 35.58%,
PARTIAL_REPEALS ≈ 11.42%, REPEALS ≈ 2.03%.

In PageRank, edge weights scale how much "influence mass" flows along each outgoing edge
before per-source normalization. High weights at a source skew more mass to those edge types;
row normalization still divides a node's total mass across its outgoing edges.

Weighted variants (edge types: CITES, AMENDS, PARTIAL_REPEALS, REPEALS):
  1) Authority_Legal_Dependence:      0.20, 1.00, 0.70, 0.05
     legal_rationale:
       - An amendment operationally edits the target Act; it expresses strong legal dependence on its definitions/structure.
       - Partial repeals still rely on the target’s scaffolding but narrow its scope; weaker than an amend.
       - Citations are heterogeneous (background, definitional, procedural); treat as weak endorsements.
       - Repeals typically signal displacement of authority; assign near-zero toward present authority.
     centrality_interpretation:
       - Influence accrues primarily to Acts that are continuously refined by later instruments (amend-heavy hubs).
       - Acts frequently tweaked/maintained accumulate more PageRank since many high-weight edges point into them.
     expected_rank_shifts:
       - Up: consolidation/“principal” Acts (frequent targets of AMENDS/PARTIAL_REPEALS).
       - Down: Acts with many bare CITES only; repealed targets do not gain much.
     when_to_use: “Which Acts are most legally foundational today?”
     caveats:
       - Under-weights landmark but heavily cited Acts with few formal changes.
     diagnostics:
       - Compare with unweighted PR: check Spearman/Kendall on full ranking and Top-k overlap (k ∈ {20, 50, 100}).
       - Check inbound composition of Top-k: expect AMENDS/PARTIAL share ↑ versus baseline.
     robustness:
       - ±10–20% jitter on weights; bootstrapped Top-k stability.
       - Ablate CITES (set to 0) to confirm authority stems from structural-change edges.
  (Only the above variant is included in program output.)

Outputs five sheets (unweighted + 1 variant + utilities) with columns:
  Act, centrality, status

Writes to: outputs/analytics/page_rank_YYYYMMDD_HHMMSS.xlsx

Cross-cutting implementation notes
- edge_direction:
    default: "X → Y when X cites/amends/partially repeals/repeals Y; influence flows to Y (the target)."
    variant: "If you conceptualize the agent of change as influential, reverse direction for AMENDS/PARTIAL/REPEALS only."
- normalization:
    - Use per-source row-normalization (standard in PageRank). Weights act as pre-normalization multipliers.
    - Watch prolific out-degree nodes (mega-bills); their mass divides over many edges even with high weights.
- temporal:
    - Compute PR in rolling windows or pre-/post- consolidation years to separate historical vs current authority.
    - Set REPEALS weight higher for historical snapshots; set to 0 for present-day authority.
- evaluation:
    - Report Spearman/Kendall with unweighted PR, Top-k overlap, and overlap with expert/curated lists if available.
    - Do weight-perturbation sweeps and ablations (set one weight to 0) to document stability.
- data_linkage:
    - If the pipeline aggregates multi-edges (e.g., multiple citations), sum counts before applying weights; or treat each edge instance with the same weight—keep consistent across runs.

Notes
- Relationship-type synonyms are supported (CITES/CIT, AMENDS/AMD, PARTIAL_REPEALS/PRP, REPEALS/FRP);
  the canonical weights are expanded to whichever names exist in the DB.
- For each weighted run, the relationship property `weight` is set on existing edges
  to the values from the respective set (idempotent overwrite per run).
"""

from __future__ import annotations

import argparse
import math
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple, Set

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series

from src.graph.neo4j_connection import close_driver, neo4j_session





# Relationship types considered for this analysis, with synonyms that may appear
# in the database. We will expand weights across whichever of these types exist.
SYNONYMS: Mapping[str, Sequence[str]] = {
    "CITES": ("CITES", "CIT"),
    "AMENDS": ("AMENDS", "AMD"),
    "PARTIAL_REPEALS": ("PARTIAL_REPEALS", "PRP"),
    "REPEALS": ("REPEALS", "FRP"),
}

# Flat set of all supported relationship types (for discovery against the DB)
SUPPORTED_REL_TYPES: Set[str] = {alias for aliases in SYNONYMS.values() for alias in aliases}


# Two distinct weight sets to run, keyed by short, sheet-friendly names.
# Values are specified by canonical edge type; synonyms get expanded automatically.
# Each set includes a short comment to explain the combination.
WEIGHT_SETS: Mapping[str, Mapping[str, float]] = {
    # Unweighted (uniform) — identical to classic unweighted PageRank when no
    # parallel same-type relations exist between a pair of Acts.
    # rationale: Treat all relation types equally and let row-normalization
    # divide mass evenly across outgoing edges.
    "unweighted": {
        "CITES": 1.0,
        "AMENDS": 1.0,
        "PARTIAL_REPEALS": 1.0,
        "REPEALS": 1.0,
    },
    # Authority_Legal_Dependence
    # legal_rationale:
    #   - Edits (AMENDS) imply strong dependence; partial repeals retain scaffolding; citations are weak; repeals near-zero today.
    # centrality_interpretation:
    #   - Acts continuously refined by later instruments accrue influence.
    # expected_rank_shifts: Up principal Acts; Down bare-citation Acts; repealed targets gain little.
    # when_to_use: Identify today’s legally foundational Acts.
    # caveats: Under-weights landmark but sparsely amended Acts.
    # diagnostics: Compare with unweighted PR; check Top-k inbound composition.
    # robustness: Jitter ±10–20%; ablate CITES to confirm dependence on structural-change edges.
    "authority_legal_dependence": {
        "CITES": 0.20,
        "AMENDS": 1.00,
        "PARTIAL_REPEALS": 0.70,
        "REPEALS": 0.05,
    },

    # (Other experimental weightings removed to simplify analysis and reporting.)
}


@dataclass(frozen=True)
class Act:
    title: str
    status: str


def _drop_graph(session, graph_name: str) -> None:
    """Drop an existing GDS graph if it exists."""
    try:
        record = (
            session.run(
                """
                CALL gds.graph.exists($graph_name)
                YIELD exists
                RETURN exists
                """,
                graph_name=graph_name,
            ).single()
            or {}
        )
    except Exception:
        return

    if record.get("exists"):
        session.run(
            """
            CALL gds.graph.drop($graph_name, false)
            YIELD graphName
            RETURN graphName
            """,
            graph_name=graph_name,
        )


def _project_unweighted_graph(session, graph_name: str, rel_types: Sequence[str]) -> None:
    """Project an unweighted directed graph of Act nodes and selected relationships."""
    _drop_graph(session, graph_name)
    session.run(
        """
        CALL gds.graph.project(
            $graph_name,
            ['Act'],
            $rel_types
        )
        """,
        graph_name=graph_name,
        rel_types=rel_types,
    )


def _project_weighted_graph(session, graph_name: str, rel_types: Sequence[str]) -> None:
    """Project a weighted directed graph of Act nodes with weights by relationship type.

    Uses native projection with property aggregation (SUM) over parallel relationships.
    Assumes a `weight` property exists on stored relationships (seeded beforehand).
    """
    _drop_graph(session, graph_name)

    rel_proj = {
        rel_type: {
            "type": rel_type,
            "orientation": "NATURAL",
            "properties": {"weight": {"property": "weight", "aggregation": "SUM"}},
        }
        for rel_type in rel_types
    }

    session.run(
        """
        CALL gds.graph.project(
            $graph_name,
            ['Act'],
            $rel_proj
        )
        """,
        graph_name=graph_name,
        rel_proj=rel_proj,
    )


def _stream_pagerank(
    session,
    *,
    graph_name: str,
    damping: float,
    max_iter: int,
    tol: float,
    weight_property: Optional[str] = None,
) -> pd.DataFrame:
    """Run PageRank in stream mode and return a DataFrame with Act, centrality, status."""
    config_parts = [
        f"dampingFactor: {float(damping)}",
        f"maxIterations: {int(max_iter)}",
        f"tolerance: {float(tol)}",
    ]
    if weight_property:
        config_parts.append(f"relationshipWeightProperty: '{weight_property}'")
    config = "{" + ", ".join(config_parts) + "}"

    query = f"""
        CALL gds.pageRank.stream($graph_name, {config})
        YIELD nodeId, score
        RETURN
            gds.util.asNode(nodeId).title AS Act,
            coalesce(gds.util.asNode(nodeId).status, 'unknown') AS status,
            score AS centrality
        ORDER BY centrality DESC, Act ASC
    """
    records = session.run(query, graph_name=graph_name)
    rows = records.data()
    return pd.DataFrame(rows, columns=["Act", "status", "centrality"])



def _compute_unweighted(session, *, damping: float, max_iter: int, tol: float, rel_types: Sequence[str]) -> pd.DataFrame:
    graph_name = "legislation_pagerank_unweighted"
    _project_unweighted_graph(session, graph_name, rel_types)
    try:
        return _stream_pagerank(
            session,
            graph_name=graph_name,
            damping=damping,
            max_iter=max_iter,
            tol=tol,
        )
    finally:
        _drop_graph(session, graph_name)


def _compute_weighted(session, *, damping: float, max_iter: int, tol: float, rel_types: Sequence[str]) -> pd.DataFrame:
    graph_name = "legislation_pagerank_weighted"
    _project_weighted_graph(session, graph_name, rel_types)
    try:
        return _stream_pagerank(
            session,
            graph_name=graph_name,
            damping=damping,
            max_iter=max_iter,
            tol=tol,
            weight_property="weight",
        )
    finally:
        _drop_graph(session, graph_name)


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(path)
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def run(
    damping: float = 0.85,
    max_iter: int = 200,
    tol: float = 1e-6,
    *,
    produce_figures: bool = False,
    figures_outdir: Optional[str] = None,
    embed_charts: bool = False,
    top_movers: int = 20,
    export_csv: bool = False,
) -> str:
    """Compute PageRank (uniform "unweighted" + weighted variants) using Neo4j GDS and export to Excel."""
    with neo4j_session() as session:
        # Determine which of our supported relationship types exist in the DB
        present = _discover_present_types(session)
        if not present:
            raise RuntimeError("No supported relationship types found for PageRank.")

        # Print the data source size once (nodes and edges among Acts for present relationship types)
        _print_source_counts(session, sorted(present))

        # For each weight set (including "unweighted"=uniform), seed weights on present
        # relationship types and compute weighted PageRank
        weighted_results: Dict[str, pd.DataFrame] = {}
        for key, canonical_weights in WEIGHT_SETS.items():
            expanded = _expand_weights_for_present_types(canonical_weights, present)
            if not expanded:
                # Skip if none of the types are present in the database
                continue
            print(f"Projecting graph and running weighted PageRank (GDS) for set: {key} ...")
            _seed_weight_properties(session, expanded)
            df_w = _compute_weighted(
                session,
                damping=damping,
                max_iter=max_iter,
                tol=tol,
                rel_types=sorted(expanded.keys()),
            )
            weighted_results[key] = df_w

        # Also compute incoming relation counts (how many relations of each type each Act received)
        print("Computing incoming relation counts per Act...")
        df_incoming = _incoming_relation_counts(session)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join("outputs", "analytics", f"page_rank_{ts}.xlsx")
    ensure_parent_dir(out_path)

    # Prepare optional comparison frames (only if both sheets exist)
    comp_df = None
    metrics_df = None
    if "unweighted" in weighted_results and "authority_legal_dependence" in weighted_results:
        comp_df = _build_comparison_table(
            weighted_results["unweighted"], weighted_results["authority_legal_dependence"]
        )
        metrics_df = _build_metrics_table(
            weighted_results["unweighted"],
            weighted_results["authority_legal_dependence"],
            ks=[10, 20, 50, 100],
        )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Ensure column order and names are consistent with previous output
        # Write all weight-set runs (including the "unweighted" uniform set)
        for key, df in weighted_results.items():
            df_out = df[["Act", "centrality", "status"]]
            df_out.to_excel(writer, sheet_name=key, index=False)
            _autofit_columns(writer, key, df_out)
        # Add incoming relation counts sheet
        df_incoming.to_excel(writer, sheet_name="incoming_relations", index=False)
        _autofit_columns(writer, "incoming_relations", df_incoming)

        # Add pairwise comparison and metrics for thesis-ready reporting (if available)
        if comp_df is not None:
            comp_df.to_excel(
                writer,
                sheet_name="compare_unweighted_vs_authority",
                index=False,
            )
            _autofit_columns(writer, "compare_unweighted_vs_authority", comp_df)
        if metrics_df is not None:
            metrics_df.to_excel(writer, sheet_name="comparison_metrics", index=False)
            _autofit_columns(writer, "comparison_metrics", metrics_df)

        # Optionally embed charts into this workbook
        if embed_charts:
            try:
                _embed_charts_in_workbook(
                    writer,
                    weighted_results=weighted_results,
                    df_incoming=df_incoming,
                    comp_df=comp_df,
                    metrics_df=metrics_df,
                    top_movers=top_movers,
                )
            except Exception:
                # Non-fatal; charts are optional
                pass

    print(f"Wrote PageRank results to {out_path}")
    # Optionally export each sheet as a standalone CSV next to the Excel workbook
    if export_csv:
        base_no_ext, _ = os.path.splitext(out_path)
        # Weighted result sheets
        for key, df in weighted_results.items():
            try:
                df_out = df[["Act", "centrality", "status"]]
                csv_path = f"{base_no_ext}__{key}.csv"
                df_out.to_csv(csv_path, index=False)
            except Exception:
                pass
        # Incoming relations
        try:
            csv_path = f"{base_no_ext}__incoming_relations.csv"
            df_incoming.to_csv(csv_path, index=False)
        except Exception:
            pass
        # Comparisons (if available)
        if comp_df is not None:
            try:
                csv_path = f"{base_no_ext}__compare_unweighted_vs_authority.csv"
                comp_df.to_csv(csv_path, index=False)
            except Exception:
                pass
        if metrics_df is not None:
            try:
                csv_path = f"{base_no_ext}__comparison_metrics.csv"
                metrics_df.to_csv(csv_path, index=False)
            except Exception:
                pass
    # Optionally produce separate PNG figures in the same folder as the Excel (default)
    if produce_figures:
        try:
            outdir = figures_outdir or os.path.dirname(out_path)
            _generate_figures_png(
                weighted_results=weighted_results,
                df_incoming=df_incoming,
                comp_df=comp_df,
                metrics_df=metrics_df,
                outdir=outdir,
                top_movers=top_movers,
            )
            print(f"Wrote figures to {outdir}")
        except Exception:
            # Non-fatal; figures are optional
            print("Failed to produce PNG figures (optional)")
    return out_path


def _discover_present_types(session) -> Set[str]:
    """Return the subset of supported relationship types that exist in the DB."""
    result = session.run(
        """
        MATCH ()-[r]->()
        RETURN DISTINCT type(r) AS type
        """
    )
    types_in_db = {rec["type"] for rec in result if rec and rec.get("type")}
    return {t for t in SUPPORTED_REL_TYPES if t in types_in_db}


def _print_source_counts(session, rel_types: Sequence[str]) -> None:
    """Print counts of Act nodes and Act-to-Act edges for the given relationship types.

    Printed once since the data source is identical across all centrality runs.
    """
    try:
        rec_n = (
            session.run(
                """
                MATCH (a:Act)
                RETURN count(a) AS n
                """
            ).single()
            or {}
        )
        n = rec_n.get("n", 0)

        rec_m = (
            session.run(
                """
                MATCH (a:Act)-[r]->(b:Act)
                WHERE type(r) IN $types
                RETURN count(r) AS m
                """,
                types=list(rel_types),
            ).single()
            or {}
        )
        m = rec_m.get("m", 0)
        print(f"Centrality data source: nodes={n}, edges={m}")
    except Exception:
        # Best-effort logging; do not fail run if counting is unavailable
        pass


def _seed_weight_properties(session, weights_by_type: Mapping[str, float]) -> None:
    """Set a numeric relationship property `weight` by relationship type.

    Accepts a mapping of actual DB relationship type names -> weight. Idempotent over-writes.
    """
    for rel_type, weight in weights_by_type.items():
        session.run(
            f"""
            MATCH ()-[r:`{rel_type}`]->()
            SET r.weight = $w
            """,
            w=float(weight),
        )


def _expand_weights_for_present_types(
    canonical_weights: Mapping[str, float], present_types: Set[str]
) -> Dict[str, float]:
    """Expand canonical weights onto whichever synonym types exist in the DB.

    Example: if only `CIT` (but not `CITES`) exists, then the weight for canonical
    `CITES` is assigned to `CIT`.
    """
    expanded: Dict[str, float] = {}
    for canonical, weight in canonical_weights.items():
        for alias in SYNONYMS.get(canonical, (canonical,)):
            if alias in present_types:
                expanded[alias] = float(weight)
    return expanded


def _incoming_relation_counts(session) -> pd.DataFrame:
    """Return counts of incoming relationships per Act by canonical type.

    Columns: Act, status, total_incoming, CITES, AMENDS, PARTIAL_REPEALS, REPEALS.
    Synonym types are mapped onto canonical columns.
    """
    present = _discover_present_types(session)

    # Build parameter lists for each canonical column using only types present.
    params = {
        "CITES": [t for t in SYNONYMS["CITES"] if t in present],
        "AMENDS": [t for t in SYNONYMS["AMENDS"] if t in present],
        "PARTIAL_REPEALS": [t for t in SYNONYMS["PARTIAL_REPEALS"] if t in present],
        "REPEALS": [t for t in SYNONYMS["REPEALS"] if t in present],
    }

    query = """
        MATCH (a:Act)
        OPTIONAL MATCH (:Act)-[r]->(a)
        WITH a, type(r) AS t
        RETURN
            a.title AS Act,
            coalesce(a.status, 'unknown') AS status,
            sum(CASE WHEN t IN $CITES THEN 1 ELSE 0 END) AS CITES,
            sum(CASE WHEN t IN $AMENDS THEN 1 ELSE 0 END) AS AMENDS,
            sum(CASE WHEN t IN $PARTIAL_REPEALS THEN 1 ELSE 0 END) AS PARTIAL_REPEALS,
            sum(CASE WHEN t IN $REPEALS THEN 1 ELSE 0 END) AS REPEALS
        ORDER BY Act ASC
    """

    rows = session.run(
        query,
        **params,
    ).data()

    df = pd.DataFrame(
        rows,
        columns=[
            "Act",
            "status",
            "CITES",
            "AMENDS",
            "PARTIAL_REPEALS",
            "REPEALS",
        ],
    )
    # Compute total incoming across all canonical relation types
    rel_cols = ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]
    df["total_incoming"] = df[rel_cols].sum(axis=1)
    # Reorder columns for readability and sort by total_incoming desc, Act asc
    df = df[["Act", "status", "total_incoming"] + rel_cols]
    df = df.sort_values(["total_incoming", "Act"], ascending=[False, True])
    return df


def _autofit_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Adjust column widths to fit content for the given sheet.

    Uses a simple character-count heuristic: max(len(header), len(str(cell))) + padding.
    Caps width to avoid excessively wide columns.
    """
    try:
        ws = writer.sheets.get(sheet_name)
        if ws is None:
            return
        max_width = 80
        padding = 2
        for idx, col in enumerate(df.columns, start=1):
            series = df[col]
            # Convert to string to measure; handle NaNs and None
            lengths = [len(str(col))] + [len(str(x)) for x in series.fillna("").astype(str).tolist()]
            width = min(max(lengths) + padding, max_width)
            ws.column_dimensions[get_column_letter(idx)].width = width
    except Exception:
        # Best-effort formatting; safe to ignore errors
        pass


def _build_comparison_table(df_unw: pd.DataFrame, df_auth: pd.DataFrame) -> pd.DataFrame:
    """Return a merged table comparing unweighted vs authority_legal_dependence.

    Columns: Act, status, rank_unweighted, centrality_unweighted, rank_authority, centrality_authority,
             delta_rank (unweighted - authority), delta_centrality (authority - unweighted).
    Sorted by delta_rank desc, then by Act.
    """
    def _ranked(df: pd.DataFrame) -> pd.DataFrame:
        r = df.sort_values(["centrality", "Act"], ascending=[False, True]).reset_index(drop=True)
        r["rank"] = r.index + 1
        return r[["Act", "status", "rank", "centrality"]]

    a = _ranked(df_unw).rename(columns={"rank": "rank_unweighted", "centrality": "centrality_unweighted"})
    b = _ranked(df_auth).rename(columns={"rank": "rank_authority", "centrality": "centrality_authority"})
    merged = pd.merge(a, b[["Act", "rank_authority", "centrality_authority"]], on="Act", how="inner")
    merged["delta_rank"] = merged["rank_unweighted"] - merged["rank_authority"]
    merged["delta_centrality"] = merged["centrality_authority"] - merged["centrality_unweighted"]
    merged = merged.sort_values(["delta_rank", "Act"], ascending=[False, True])
    return merged


def _build_metrics_table(df_unw: pd.DataFrame, df_auth: pd.DataFrame, *, ks: Sequence[int]) -> pd.DataFrame:
    """Compute simple comparison metrics between unweighted and authority_legal_dependence.

    Metrics include Spearman/Kendall correlations (full list) and Overlap@k + Jaccard@k.
    """
    # Align by Act
    a = df_unw[["Act", "centrality"]].rename(columns={"centrality": "unweighted"})
    b = df_auth[["Act", "centrality"]].rename(columns={"centrality": "authority"})
    merged = pd.merge(a, b, on="Act", how="inner")

    metrics_rows = []
    # Rank correlations
    try:
        spearman = merged[["unweighted", "authority"]].corr(method="spearman").iloc[0, 1]
        metrics_rows.append({"metric": "spearman", "k": "all", "value": float(spearman)})
    except Exception:
        pass
    try:
        kendall = merged[["unweighted", "authority"]].corr(method="kendall").iloc[0, 1]
        metrics_rows.append({"metric": "kendall", "k": "all", "value": float(kendall)})
    except Exception:
        pass

    # Top-k overlaps and Jaccards
    def _topk(df: pd.DataFrame, k: int) -> Set[str]:
        return set(
            df.sort_values(["centrality", "Act"], ascending=[False, True])
            .head(k)["Act"]
            .tolist()
        )

    for k in ks:
        try:
            sa = _topk(df_unw, k)
            sb = _topk(df_auth, k)
            inter = len(sa & sb)
            union = len(sa | sb) if (sa or sb) else 0
            jacc = (inter / union) if union else 0.0
            metrics_rows.append({"metric": "overlap@k", "k": int(k), "value": int(inter)})
            metrics_rows.append({"metric": "jaccard@k", "k": int(k), "value": float(jacc)})
        except Exception:
            continue

    return pd.DataFrame(metrics_rows, columns=["metric", "k", "value"])


def _generate_figures_png(
    *,
    weighted_results: Mapping[str, pd.DataFrame],
    df_incoming: pd.DataFrame,
    comp_df: Optional[pd.DataFrame],
    metrics_df: Optional[pd.DataFrame],
    outdir: str,
    top_movers: int = 20,
) -> None:
    # Try matplotlib; skip if not available
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        print("matplotlib not installed; skipping PNG figure generation")
        return
    os.makedirs(outdir, exist_ok=True)

    # Figure 1: Scatter unweighted vs authority
    if "unweighted" in weighted_results and "authority_legal_dependence" in weighted_results:
        _plot_scatter_unw_vs_auth_png(
            weighted_results["unweighted"],
            weighted_results["authority_legal_dependence"],
            os.path.join(outdir, "pr_scatter_unw_vs_auth.png"),
        )

    # Figure 2: Rank delta top movers
    if comp_df is not None:
        _plot_rank_delta_topk_png(
            comp_df,
            os.path.join(outdir, "pr_rank_delta_top20.png"),
            top_movers=top_movers,
        )

    # Figure 3: Incoming composition for movers
    if comp_df is not None and df_incoming is not None:
        _plot_incoming_composition_for_movers_png(
            comp_df,
            df_incoming,
            os.path.join(outdir, "pr_incoming_composition_movers.png"),
            top_movers=top_movers,
        )

    # Figure 4: Overlap/Jaccard vs k (if metrics present)
    if metrics_df is not None and not metrics_df.empty:
        _plot_overlap_jaccard_png(metrics_df, os.path.join(outdir, "pr_overlap_jaccard.png"))

    # Figure 5: Before/after leaderboard
    if "unweighted" in weighted_results and "authority_legal_dependence" in weighted_results:
        _plot_leaderboard_before_after_png(
            weighted_results["unweighted"],
            weighted_results["authority_legal_dependence"],
            os.path.join(outdir, "pr_leaderboard_before_after.png"),
            top_n=10,
        )


def _plot_scatter_unw_vs_auth_png(df_unw: pd.DataFrame, df_auth: pd.DataFrame, out_path: str) -> None:
    import matplotlib.pyplot as plt
    import pandas as pd

    a = df_unw[["Act", "centrality"]].rename(columns={"centrality": "unweighted"})
    b = df_auth[["Act", "centrality"]].rename(columns={"centrality": "authority"})
    df = pd.merge(a, b, on="Act", how="inner")

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.scatter(df["unweighted"], df["authority"], alpha=0.4, s=12, color="#5B8FF9")
    mn = min(df["unweighted"].min(), df["authority"].min())
    mx = max(df["unweighted"].max(), df["authority"].max())
    ax.plot([mn, mx], [mn, mx], color="#999999", linestyle="--", linewidth=1)
    ax.set_xlabel("Unweighted PageRank score")
    ax.set_ylabel("Authority-weighted PageRank score")
    ax.set_title("Unweighted vs Authority-Weighted PageRank")
    fig.tight_layout()
    fig.savefig(out_path, dpi=200)
    plt.close(fig)


def _plot_rank_delta_topk_png(comp_df: pd.DataFrame, out_path: str, *, top_movers: int = 20) -> None:
    import matplotlib.pyplot as plt

    comp = comp_df.sort_values(["delta_rank", "Act"], ascending=[False, True]).head(top_movers)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(comp["Act"][::-1], comp["delta_rank"][::-1], color="#5AD8A6")
    ax.set_xlabel("Rank improvement (unweighted rank - authority rank)")
    ax.set_ylabel("Act")
    ax.set_title(f"Top {top_movers} Rank Improvements under Authority Weighting")
    fig.tight_layout()
    fig.savefig(out_path, dpi=200)
    plt.close(fig)


def _plot_incoming_composition_for_movers_png(
    comp_df: pd.DataFrame,
    df_incoming: pd.DataFrame,
    out_path: str,
    *,
    top_movers: int = 20,
) -> None:
    import matplotlib.pyplot as plt
    import pandas as pd

    movers = comp_df.sort_values(["delta_rank", "Act"], ascending=[False, True]).head(top_movers)
    acts = movers["Act"].tolist()

    sub = df_incoming[df_incoming["Act"].isin(acts)].copy()
    for col in ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]:
        sub[col] = pd.to_numeric(sub[col], errors="coerce").fillna(0)
    sub["total"] = sub[["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]].sum(axis=1)
    for col in ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]:
        sub[col] = sub[col] / sub["total"].replace(0, 1)
    sub = sub.set_index("Act").loc[acts]

    fig, ax = plt.subplots(figsize=(10, max(6, 0.4 * len(acts))))
    bottom = None
    colors = {
        "CITES": "#5B8FF9",
        "AMENDS": "#5AD8A6",
        "PARTIAL_REPEALS": "#F6BD16",
        "REPEALS": "#E8684A",
    }
    for col in ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]:
        vals = sub[col]
        ax.barh(sub.index, vals, left=bottom, color=colors[col], label=col)
        bottom = vals if bottom is None else bottom + vals
    ax.legend(loc="lower right", ncol=2)
    ax.set_xlabel("Incoming composition share")
    ax.set_ylabel("Act")
    ax.set_title(f"Incoming Relation Composition for Top {top_movers} Movers")
    fig.tight_layout()
    fig.savefig(out_path, dpi=200)
    plt.close(fig)


def _plot_overlap_jaccard_png(metrics_df: pd.DataFrame, out_path: str) -> None:
    import matplotlib.pyplot as plt
    import pandas as pd

    m = metrics_df.copy()
    m = m[m["metric"].isin(["overlap@k", "jaccard@k"])].copy()
    m["k"] = pd.to_numeric(m["k"], errors="coerce")
    if m["k"].isna().all():
        return
    m = m.dropna(subset=["k"]).sort_values("k")

    fig, ax1 = plt.subplots(figsize=(7, 4))
    for metric, color in [("overlap@k", "#5B8FF9"), ("jaccard@k", "#E8684A")]:
        sub = m[m["metric"] == metric]
        ax1.plot(sub["k"], sub["value"], marker="o", label=metric, color=color)
    ax1.set_xlabel("k")
    ax1.set_title("Overlap and Jaccard vs k (Unweighted vs Authority)")
    ax1.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=200)
    plt.close(fig)


def _plot_leaderboard_before_after_png(
    df_unw: pd.DataFrame,
    df_auth: pd.DataFrame,
    out_path: str,
    *,
    top_n: int = 10,
) -> None:
    import matplotlib.pyplot as plt
    import pandas as pd

    def _top(df: pd.DataFrame) -> pd.DataFrame:
        return df.sort_values(["centrality", "Act"], ascending=[False, True]).head(top_n)[["Act", "centrality"]]

    u = _top(df_unw).rename(columns={"centrality": "Unweighted"})
    a = _top(df_auth).rename(columns={"centrality": "Authority"})
    acts = list(dict.fromkeys(u["Act"].tolist() + a["Act"].tolist()))
    u = u.set_index("Act").reindex(acts)
    a = a.set_index("Act").reindex(acts)
    df = u.join(a, how="outer").fillna(0)

    x = list(range(len(df)))
    width = 0.4
    fig, ax = plt.subplots(figsize=(10, max(5, 0.5 * len(df))))
    ax.barh([i + width / 2 for i in x], df["Unweighted"], height=width, label="Unweighted", color="#5B8FF9")
    ax.barh([i - width / 2 for i in x], df["Authority"], height=width, label="Authority", color="#5AD8A6")
    ax.set_yticks(x)
    ax.set_yticklabels(df.index.tolist())
    ax.set_xlabel("Centrality score (run-specific scale)")
    ax.set_title("Top Acts: Before vs After Weighting")
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=200)
    plt.close(fig)


def _embed_charts_in_workbook(
    writer: pd.ExcelWriter,
    *,
    weighted_results: Mapping[str, pd.DataFrame],
    df_incoming: pd.DataFrame,
    comp_df: Optional[pd.DataFrame],
    metrics_df: Optional[pd.DataFrame],
    top_movers: int,
) -> None:
    wb = writer.book

    # Scatter: unweighted vs authority
    if "unweighted" in weighted_results and "authority_legal_dependence" in weighted_results:
        ws = wb.create_sheet(title="viz_scatter")
        ws.append(["Act", "centrality_unweighted", "centrality_authority"])
        # Build rows from merged
        a = weighted_results["unweighted"][["Act", "centrality"]].rename(columns={"centrality": "u"})
        b = weighted_results["authority_legal_dependence"]["centrality"].rename("v")
        merged = a.join(b, on="Act", how="inner")
        for _, row in merged.iterrows():
            ws.append([row["Act"], float(row["u"]), float(row["v"])])
        n = ws.max_row
        chart = ScatterChart()
        chart.title = "Unweighted vs Authority"
        chart.x_axis.title = "Unweighted"
        chart.y_axis.title = "Authority"
        xvalues = Reference(ws, min_col=2, min_row=2, max_row=n)
        yvalues = Reference(ws, min_col=3, min_row=2, max_row=n)
        series = Series(yvalues, xvalues, title="Acts")
        chart.series.append(series)
        ws.add_chart(chart, "E2")

    # Rank delta top movers
    if comp_df is not None:
        ws = wb.create_sheet(title="viz_rank_delta_topN")
        comp = comp_df.sort_values(["delta_rank", "Act"], ascending=[False, True]).head(top_movers)
        ws.append(["Act", "delta_rank"])
        for _, row in comp.iterrows():
            ws.append([row["Act"], int(row["delta_rank"])])
        n = ws.max_row
        cat = Reference(ws, min_col=1, min_row=2, max_row=n)
        val = Reference(ws, min_col=2, min_row=1, max_row=n)
        chart = BarChart()
        chart.type = "bar"  # horizontal bars
        chart.title = f"Top {top_movers} Rank Improvements"
        chart.y_axis.title = "Rank improvement"
        chart.add_data(val, titles_from_data=True)
        chart.set_categories(cat)
        ws.add_chart(chart, "D2")

    # Composition for movers (stacked)
    if comp_df is not None and df_incoming is not None:
        ws = wb.create_sheet(title="viz_composition_topN")
        comp = comp_df.sort_values(["delta_rank", "Act"], ascending=[False, True]).head(top_movers)
        acts = comp["Act"].tolist()
        sub = df_incoming[df_incoming["Act"].isin(acts)].copy()
        for col in ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]:
            sub[col] = pd.to_numeric(sub[col], errors="coerce").fillna(0)
        sub["total"] = sub[["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]].sum(axis=1)
        for col in ["CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"]:
            sub[col] = sub[col] / sub["total"].replace(0, 1)
        sub = sub.set_index("Act").loc[acts]
        ws.append(["Act", "CITES", "AMENDS", "PARTIAL_REPEALS", "REPEALS"])
        for act, r in sub.iterrows():
            ws.append([act, float(r["CITES"]), float(r["AMENDS"]), float(r["PARTIAL_REPEALS"]), float(r["REPEALS"])])
        n = ws.max_row
        data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=n)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = f"Incoming Composition Shares (Top {top_movers} Movers)"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")

    # Overlap/Jaccard vs k
    if metrics_df is not None and not metrics_df.empty:
        m = metrics_df.copy()
        m = m[m["metric"].isin(["overlap@k", "jaccard@k"])].copy()
        if not m.empty:
            ws = wb.create_sheet(title="viz_overlap_jaccard")
            ws.append(["k", "overlap@k", "jaccard@k"])
            # Pivot rows by k
            ks = sorted({int(x) for x in m["k"] if str(x).isdigit()})
            for k in ks:
                ov = m[(m["metric"] == "overlap@k") & (m["k"] == k)]["value"].astype(float)
                ja = m[(m["metric"] == "jaccard@k") & (m["k"] == k)]["value"].astype(float)
                ovv = float(ov.iloc[0]) if not ov.empty else 0.0
                jav = float(ja.iloc[0]) if not ja.empty else 0.0
                ws.append([k, ovv, jav])
            n = ws.max_row
            data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=n)
            cats = Reference(ws, min_col=1, min_row=2, max_row=n)
            chart = LineChart()
            chart.title = "Overlap and Jaccard vs k"
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "E2")

    # Before/after leaderboard (union of Top 10s)
    if "unweighted" in weighted_results and "authority_legal_dependence" in weighted_results:
        ws = wb.create_sheet(title="viz_leaderboard")
        def _top(df: pd.DataFrame) -> pd.DataFrame:
            return df.sort_values(["centrality", "Act"], ascending=[False, True]).head(10)[["Act", "centrality"]]
        u = _top(weighted_results["unweighted"]).rename(columns={"centrality": "Unweighted"})
        a = _top(weighted_results["authority_legal_dependence"]).rename(columns={"centrality": "Authority"})
        acts = list(dict.fromkeys(u["Act"].tolist() + a["Act"].tolist()))
        u = u.set_index("Act").reindex(acts)
        a = a.set_index("Act").reindex(acts)
        ws.append(["Act", "Unweighted", "Authority"])
        for act in acts:
            ws.append([act, float(u.loc[act, "Unweighted"]) if act in u.index and pd.notna(u.loc[act, "Unweighted"]) else 0.0,
                            float(a.loc[act, "Authority"]) if act in a.index and pd.notna(a.loc[act, "Authority"]) else 0.0])
        n = ws.max_row
        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=n)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Top Acts: Before vs After Weighting"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compute PageRank (unweighted & weighted) for Acts in Neo4j.")
    parser.add_argument("--damping", type=float, default=0.85, help="Damping factor (default: 0.85)")
    parser.add_argument("--max-iter", type=int, default=200, help="Maximum iterations (default: 200)")
    parser.add_argument("--tol", type=float, default=1e-6, help="Convergence tolerance (default: 1e-6)")
    parser.add_argument("--figures", action="store_true", help="Also generate PNG figures next to the Excel (default dir = Excel folder)")
    parser.add_argument("--figures-outdir", default=None, help="Directory to write figures (defaults to Excel folder)")
    parser.add_argument("--embed-charts", action="store_true", help="Embed charts as Excel sheets instead of PNGs")
    parser.add_argument("--top-movers", type=int, default=20, help="Number of top movers to visualize (default: 20)")
    parser.add_argument("--export-csv", action="store_true", help="Also export each Excel sheet as a standalone CSV next to the workbook")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        run(
            damping=args.damping,
            max_iter=args.max_iter,
            tol=args.tol,
            produce_figures=args.figures,
            figures_outdir=args.figures_outdir,
            embed_charts=args.embed_charts,
            top_movers=args.top_movers,
            export_csv=args.export_csv,
        )
    except Exception:
        print("Failed to compute PageRank")
        raise
    finally:
        close_driver()


if __name__ == "__main__":
    main()
